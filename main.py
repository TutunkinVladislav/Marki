import openpyxl as op


filename = r'c:\Users\pocht\OneDrive\Рабочий стол\Марки.xlsx'
list_marka = {}

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active

max_rows = sheet.max_row

for i in range(1, max_rows+1):
    name = sheet.cell(row=i, column=1).value
    marka = sheet.cell(row=i, column=2).value
    if name not in list_marka:
        list_marka[name] = [marka]
    else:
        list_marka[name].append(marka)

with open('marki.txt', 'w') as file:
    for key, value in list_marka.items():
        string_marki = ', '.join(value)
        string_to_write = key + ' = ' + string_marki + '\n'
        file.write(string_to_write)
